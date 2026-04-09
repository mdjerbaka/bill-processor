"""Payables tracking and Excel export service."""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select, case
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.models.models import (
    AppSetting,
    Invoice,
    InvoiceStatus,
    Job,
    Payable,
    PayableStatus,
)

logger = logging.getLogger(__name__)


class PayablesService:
    """Manages payables tracking and exports."""

    def __init__(self, db: AsyncSession, user_id: int):
        self.db = db
        self.user_id = user_id

    async def create_payable(self, invoice: Invoice) -> Payable:
        """Create a payable record from an approved invoice."""
        # Get job name if invoice has a matched job
        job_name = None
        if invoice.job:
            job_name = invoice.job.name

        payable = Payable(
            invoice_id=invoice.id,
            vendor_name=invoice.vendor_name or "Unknown",
            amount=invoice.total_amount or 0.0,
            due_date=invoice.due_date,
            status=PayableStatus.OUTSTANDING,
            user_id=self.user_id,
            invoice_number=invoice.invoice_number,
            job_name=job_name,
            qbo_bill_id=invoice.qbo_bill_id,
            qbo_vendor_id=invoice.qbo_vendor_id,
        )
        self.db.add(payable)
        await self.db.flush()
        return payable

    async def get_outstanding_payables(self) -> list[Payable]:
        """Get all outstanding (unpaid) payables."""
        result = await self.db.execute(
            select(Payable)
            .where(
                Payable.user_id == self.user_id,
                Payable.status.in_([PayableStatus.OUTSTANDING, PayableStatus.OVERDUE]),
            )
            .order_by(Payable.due_date.asc())
        )
        return list(result.scalars().all())

    async def get_payables_summary(self) -> dict:
        """Get aggregated payables summary."""
        now = datetime.now(timezone.utc)

        # First, update overdue status
        await self.db.execute(
            Payable.__table__.update()
            .where(
                Payable.user_id == self.user_id,
                Payable.status == PayableStatus.OUTSTANDING,
                Payable.due_date < now,
                Payable.due_date.is_not(None),
            )
            .values(status=PayableStatus.OVERDUE)
        )
        await self.db.flush()

        # Base filter: non-junked, non-permanent, active payables for this user
        base_filter = [
            Payable.user_id == self.user_id,
            Payable.is_junked == False,  # noqa: E712
            Payable.is_permanent == False,  # noqa: E712
            Payable.status.in_([PayableStatus.OUTSTANDING, PayableStatus.OVERDUE]),
        ]

        # Total outstanding (all non-junked, non-permanent)
        result = await self.db.execute(
            select(func.coalesce(func.sum(Payable.amount), 0.0)).where(*base_filter)
        )
        total_outstanding = result.scalar()

        # Total overdue (all non-junked, non-permanent)
        result = await self.db.execute(
            select(func.coalesce(func.sum(Payable.amount), 0.0)).where(
                *base_filter,
                Payable.status == PayableStatus.OVERDUE,
            )
        )
        total_overdue = result.scalar()

        # Count
        result = await self.db.execute(
            select(func.count(Payable.id)).where(*base_filter)
        )
        count = result.scalar()

        return {
            "total_outstanding": float(total_outstanding),
            "total_overdue": float(total_overdue),
            "count": count,
        }

    async def get_real_balance(self) -> dict:
        """Calculate real available funds.

        Formula: bank + receivables (collect) - outstanding payables - buffer
        """
        # Get current bank balance from settings
        result = await self.db.execute(
            select(AppSetting).where(AppSetting.key == "bank_balance", AppSetting.user_id == self.user_id)
        )
        setting = result.scalar_one_or_none()
        bank_balance = float(setting.value) if setting else 0.0

        # Get buffer amount
        result = await self.db.execute(
            select(AppSetting).where(AppSetting.key == "balance_buffer", AppSetting.user_id == self.user_id)
        )
        buf_setting = result.scalar_one_or_none()
        buffer = float(buf_setting.value) if buf_setting else 0.0

        summary = await self.get_payables_summary()

        # Get total receivables expected to collect
        from app.models.models import ReceivableCheck
        recv_result = await self.db.execute(
            select(func.coalesce(func.sum(ReceivableCheck.invoiced_amount), 0.0))
            .where(
                ReceivableCheck.user_id == self.user_id,
                ReceivableCheck.collect == True,  # noqa: E712
            )
        )
        total_receivables = float(recv_result.scalar() or 0.0)

        return {
            "bank_balance": bank_balance,
            "total_outstanding": summary["total_outstanding"],
            "total_receivables": total_receivables,
            "buffer": buffer,
            "real_available": bank_balance + total_receivables - summary["total_outstanding"] - buffer,
        }

    async def export_to_excel(self) -> bytes:
        """Export payables to an Excel spreadsheet matching 'Real Bank Balance' format."""
        # Fetch non-junked, non-permanent payables with optional invoice info
        result = await self.db.execute(
            select(Payable, Invoice, Job)
            .outerjoin(Invoice, Payable.invoice_id == Invoice.id)
            .outerjoin(Job, Invoice.job_id == Job.id)
            .where(
                Payable.user_id == self.user_id,
                Payable.is_junked == False,  # noqa: E712
                Payable.is_permanent == False,  # noqa: E712
                Payable.status.in_([PayableStatus.OUTSTANDING, PayableStatus.OVERDUE]),
            )
            .order_by(Payable.due_date.asc())
        )
        rows = result.all()

        # Get bank balance
        balance_data = await self.get_real_balance()

        wb = Workbook()

        # ── Payables Sheet ───────────────────────────────
        ws = wb.active
        ws.title = "Payables"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        money_format = '#,##0.00'
        date_format = 'MM/DD/YYYY'
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        overdue_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Headers
        headers = ["Vendor", "Invoice #", "Amount", "Due Date", "Status", "Job", "Days Until Due"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        now = datetime.now(timezone.utc)
        for row_idx, (payable, invoice, job) in enumerate(rows, 2):
            days_until = (payable.due_date - now).days if payable.due_date else None

            ws.cell(row=row_idx, column=1, value=payable.vendor_name).border = thin_border
            ws.cell(row=row_idx, column=2, value=payable.invoice_number or (invoice.invoice_number if invoice else "") or "").border = thin_border

            amount_cell = ws.cell(row=row_idx, column=3, value=payable.amount)
            amount_cell.number_format = money_format
            amount_cell.border = thin_border

            date_cell = ws.cell(
                row=row_idx, column=4,
                value=payable.due_date.replace(tzinfo=None) if payable.due_date else None
            )
            date_cell.number_format = date_format
            date_cell.border = thin_border

            status_cell = ws.cell(row=row_idx, column=5, value=payable.status.value.upper())
            status_cell.border = thin_border
            if payable.status == PayableStatus.OVERDUE:
                for c in range(1, 8):
                    ws.cell(row=row_idx, column=c).fill = overdue_fill

            ws.cell(row=row_idx, column=6, value=job.name if job else "").border = thin_border

            days_cell = ws.cell(row=row_idx, column=7, value=days_until)
            days_cell.border = thin_border

        # Column widths
        col_widths = [25, 15, 12, 14, 12, 25, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Summary section ──────────────────────────────
        summary_row = len(rows) + 3
        ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)

        ws.cell(row=summary_row + 1, column=1, value="Bank Balance")
        bal_cell = ws.cell(row=summary_row + 1, column=2, value=balance_data["bank_balance"])
        bal_cell.number_format = money_format

        ws.cell(row=summary_row + 2, column=1, value="Receivables (Collect)")
        recv_cell = ws.cell(row=summary_row + 2, column=2, value=balance_data["total_receivables"])
        recv_cell.number_format = money_format

        ws.cell(row=summary_row + 3, column=1, value="Total Outstanding Payables")
        out_cell = ws.cell(row=summary_row + 3, column=2, value=balance_data["total_outstanding"])
        out_cell.number_format = money_format

        ws.cell(row=summary_row + 4, column=1, value="Real Available Funds").font = Font(bold=True)
        real_cell = ws.cell(row=summary_row + 4, column=2, value=balance_data["real_available"])
        real_cell.number_format = money_format
        real_cell.font = Font(bold=True, size=12)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
